#!/usr/bin/env python3
"""读取附件 xlsx，并把"碎片"还原成母介质 A。

附件每一行不是一根完整的介质 A，而是边界截断之后的一个**碎片**：
组 3 有 535 行，但按轴向分组后只有 354 个不同方向，正好对应 354 根介质 A
（体积分数 0.50%）。数据审计（见 audit_attachment.py）还发现附件系统性地
丢弃了长度小于约 500 nm 的短碎片，因此各组碎片总长略小于 根数×5000。
"""

from __future__ import annotations

import math
from collections import defaultdict
from pathlib import Path

import numpy as np
import openpyxl

from microstructure import EDGE

ROOT = Path(__file__).resolve().parents[1]
ATTACHMENT = ROOT.parent / "01_题目与附件" / "附件.xlsx"

# 组 1、组 2 的 y、z 坐标在 ±500 处发生周期回绕（见 audit_attachment.py），
# 说明这两个微构体的截面是 1000×1000 nm，只有组 3 是题面描述的 10000 立方体。
GROUP_BOX = {
    "组1": np.array([EDGE, 1000.0, 1000.0]),
    "组2": np.array([EDGE, 1000.0, 1000.0]),
    "组3": np.array([EDGE, EDGE, EDGE]),
}


def load_pieces(path: Path | None = None) -> dict[str, np.ndarray]:
    """返回 {组名: (n,6) 数组}，每行是一个碎片的两个轴端点。"""
    wb = openpyxl.load_workbook(path or ATTACHMENT, read_only=True, data_only=True)
    out: dict[str, np.ndarray] = {}
    for name in wb.sheetnames:
        rows = [r for r in wb[name].iter_rows(values_only=True)][2:]
        vals = [[float(v) for v in r[:6]] for r in rows if r[0] is not None]
        out[name] = np.array(vals)
    return out


def _canonical_dir(p: np.ndarray, q: np.ndarray) -> tuple:
    d = q - p
    d = d / np.linalg.norm(d)
    if d[0] < 0 or (d[0] == 0 and d[1] < 0):
        d = -d
    return tuple(np.round(d, 7))


def group_by_medium(pieces: np.ndarray) -> list[list[int]]:
    """按轴向把碎片归到母介质。同一根介质的所有碎片方向完全相同。"""
    groups: dict[tuple, list[int]] = defaultdict(list)
    for i, row in enumerate(pieces):
        groups[_canonical_dir(row[:3], row[3:])].append(i)
    return list(groups.values())


def chain_pieces(pieces: np.ndarray, idx: list[int], box: np.ndarray) -> list[int]:
    """把一根母介质的碎片按首尾相接顺序排好（用于还原未截断的轴线段）。"""
    half = box / 2.0

    def on_boundary(pt):
        return [j for j in range(3) if abs(abs(pt[j]) - half[j]) < 1e-6]

    succ, pred = {}, {}
    for i in idx:
        q = pieces[i][3:]
        for j in on_boundary(q):
            tgt = q.copy()
            tgt[j] = -q[j]
            for k in idx:
                if k != i and np.allclose(pieces[k][:3], tgt, atol=1e-6):
                    succ[i], pred[k] = k, i
    heads = [i for i in idx if i not in pred]
    order = []
    for h in heads:
        cur = h
        while cur is not None and cur not in order:
            order.append(cur)
            cur = succ.get(cur)
    for i in idx:            # 兜底：链断裂时也不丢碎片
        if i not in order:
            order.append(i)
    return order


def reconstruct_axis(pieces: np.ndarray, idx: list[int], box: np.ndarray) -> tuple[np.ndarray, np.ndarray, float]:
    """还原母介质未经截断的轴线段（起点、终点、已知总长）。"""
    order = chain_pieces(pieces, idx, box)
    start = pieces[order[0]][:3].copy()
    total = sum(float(np.linalg.norm(pieces[i][3:] - pieces[i][:3])) for i in order)
    d = pieces[order[0]][3:] - pieces[order[0]][:3]
    u = d / np.linalg.norm(d)
    return start, start + total * u, total


def summarize(path: Path | None = None) -> dict[str, dict]:
    """每组的碎片数、母介质数、总长与体积分数。"""
    data = load_pieces(path)
    out = {}
    for name, pieces in data.items():
        box = GROUP_BOX[name]
        media = group_by_medium(pieces)
        lens = [float(np.linalg.norm(r[3:] - r[:3])) for r in pieces]
        out[name] = {
            "n_pieces": len(pieces),
            "n_media": len(media),
            "total_axis_length": sum(lens),
            "implied_media_by_length": sum(lens) / 5000.0,
            "box": box.tolist(),
            "missing_length": len(media) * 5000.0 - sum(lens),
        }
    return out


if __name__ == "__main__":
    import json
    print(json.dumps(summarize(), ensure_ascii=False, indent=2))
